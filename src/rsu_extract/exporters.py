git push -u origin main
from __future__ import annotations

from collections import OrderedDict
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from rsu_extract.models import SaleLotRow


@dataclass(slots=True)
class TransactionGroup:
    transaction_id: str
    sale_date: str
    settlement_date: str
    sale_price_usd: Decimal | None
    proceeds_currency: str
    converted_net_proceeds: Decimal | None
    rows: list[SaleLotRow]

    @property
    def total_shares(self) -> Decimal:
        return sum((row.shares_sold or Decimal("0")) for row in self.rows)

    @property
    def gross_proceeds_usd(self) -> Decimal:
        return sum((row.total_sale_proceeds_usd or Decimal("0")) for row in self.rows)

    @property
    def fees_usd(self) -> Decimal | None:
        if self.rows and self.rows[0].fees_usd is not None:
            return self.rows[0].fees_usd
        return None

    @property
    def commission_usd(self) -> Decimal | None:
        if self.rows and self.rows[0].commission_usd is not None:
            return self.rows[0].commission_usd
        return None

    @property
    def total_fees_usd(self) -> Decimal | None:
        values = [value for value in (self.fees_usd, self.commission_usd) if value is not None]
        if not values:
            return None
        return sum(values, Decimal("0"))

    @property
    def net_proceeds_usd(self) -> Decimal | None:
        if self.rows and self.rows[0].net_proceeds_usd is not None:
            return self.rows[0].net_proceeds_usd
        fees = self.total_fees_usd
        if fees is None:
            return None
        return self.gross_proceeds_usd - fees


def group_transactions(rows: list[SaleLotRow]) -> list[TransactionGroup]:
    grouped: OrderedDict[str, list[SaleLotRow]] = OrderedDict()
    for row in sorted(
        rows,
        key=lambda item: (
            _parse_date(item.sale_date),
            item.transaction_id,
            _parse_date(item.acquisition_date) if item.acquisition_date else datetime.max,
        ),
    ):
        key = row.transaction_id or "|".join(
            [
                row.sale_date,
                row.settlement_date,
                str(row.sale_price_usd or ""),
                str(row.commission_usd or ""),
                str(row.fees_usd or ""),
                str(row.net_proceeds_usd or ""),
            ]
        )
        grouped.setdefault(key, []).append(row)

    transactions: list[TransactionGroup] = []
    for key, transaction_rows in grouped.items():
        first = transaction_rows[0]
        transactions.append(
            TransactionGroup(
                transaction_id=key,
                sale_date=first.sale_date,
                settlement_date=first.settlement_date,
                sale_price_usd=first.sale_price_usd,
                proceeds_currency=first.proceeds_currency,
                converted_net_proceeds=first.converted_net_proceeds,
                rows=transaction_rows,
            )
        )
    return transactions


def write_xlsx(path: Path, rows: list[SaleLotRow]) -> None:
    workbook = Workbook()
    sales_sheet = workbook.active
    sales_sheet.title = "Sales Breakdown"
    transactions_sheet = workbook.create_sheet("Transaction Breakdown")

    _write_sales_breakdown(sales_sheet, rows)
    _write_transaction_breakdown(transactions_sheet, rows)
    _apply_sheet_widths(sales_sheet, 24.0, 7)
    _apply_sheet_widths(transactions_sheet, 24.0, 8)
    _style_sales_breakdown(sales_sheet)
    _style_transaction_breakdown(transactions_sheet)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _write_sales_breakdown(sheet, rows: list[SaleLotRow]) -> None:
    current_row = 1
    data_rows: list[int] = []
    for transaction in group_transactions(rows):
        sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
        title_cell = sheet.cell(row=current_row, column=1, value=transaction.sale_date)
        title_cell.alignment = Alignment(horizontal="center")
        title_cell.font = Font(size=14, bold=True)
        current_row += 1
        headers_row = current_row
        for index, header in enumerate(
            [
                "Acquisition Date",
                "Shares Sold",
                "Cost Basis per Share (USD)",
                "Sale Price per Share (USD)",
                "Total Cost Basis (USD)",
                "Total Sale Proceeds (USD)",
                "Gain/Loss (USD)",
            ],
            start=1,
        ):
            header_cell = sheet.cell(row=current_row, column=index, value=header)
            header_cell.font = Font(bold=True)
        current_row += 1

        first_data_row = current_row
        for sale_row in transaction.rows:
            sheet.cell(row=current_row, column=1, value=sale_row.acquisition_date or None)
            sheet.cell(row=current_row, column=2, value=_to_number(sale_row.shares_sold))
            sheet.cell(row=current_row, column=3, value=_to_number(sale_row.cost_basis_per_share_usd))
            sheet.cell(row=current_row, column=4, value=_to_number(sale_row.sale_price_usd))
            sheet.cell(row=current_row, column=5, value=f"=B{current_row}*C{current_row}")
            sheet.cell(row=current_row, column=6, value=f"=B{current_row}*D{current_row}")
            sheet.cell(row=current_row, column=7, value=f"=F{current_row}-E{current_row}")
            data_rows.append(current_row)
            current_row += 1

        last_data_row = current_row - 1
        sheet.cell(row=current_row, column=5, value=f"=SUM(E{first_data_row}:E{last_data_row})")
        sheet.cell(row=current_row, column=6, value=f"=SUM(F{first_data_row}:F{last_data_row})")
        sheet.cell(row=current_row, column=7, value=f"=F{current_row}-E{current_row}")
        current_row += 2

    if data_rows:
        current_row = _write_overall_totals_table(sheet, current_row, data_rows)


def _write_overall_totals_table(sheet, current_row: int, data_rows: list[int]) -> int:
    sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
    title_cell = sheet.cell(row=current_row, column=1, value="Overall Totals")
    title_cell.alignment = Alignment(horizontal="center")
    title_cell.font = Font(size=14, bold=True)
    current_row += 1

    headers = [
        "Total Shares Sold",
        "Total Cost Basis (USD)",
        "Total Sale Proceeds (USD)",
        "Total Gain/Loss (USD)",
    ]
    for index, header in enumerate(headers, start=1):
        header_cell = sheet.cell(row=current_row, column=index, value=header)
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(horizontal="center")
    current_row += 1

    share_formula = _sum_formula("B", data_rows)
    cost_formula = _sum_formula("E", data_rows)
    proceeds_formula = _sum_formula("F", data_rows)
    gain_formula = _sum_formula("G", data_rows)
    sheet.cell(row=current_row, column=1, value=share_formula)
    sheet.cell(row=current_row, column=2, value=cost_formula)
    sheet.cell(row=current_row, column=3, value=proceeds_formula)
    sheet.cell(row=current_row, column=4, value=gain_formula)
    return current_row + 1


def _write_transaction_breakdown(sheet, rows: list[SaleLotRow]) -> None:
    headers = [
        "Transaction Date",
        "Shares Sold",
        "Price Per Share",
        "Gross Proceeds USD",
        "Fees USD",
        "Net Proceeds USD",
        "Currency",
        "Converted Net Proceeds (CAD)",
    ]
    for index, header in enumerate(headers, start=1):
        header_cell = sheet.cell(row=1, column=index, value=header)
        header_cell.font = Font(bold=True)

    current_row = 2
    for transaction in group_transactions(rows):
        sale_date_cell = sheet.cell(row=current_row, column=1, value=_parse_date(transaction.sale_date))
        sale_date_cell.number_format = "yyyy-mm-dd"
        sheet.cell(row=current_row, column=2, value=_to_number(transaction.total_shares))
        sheet.cell(row=current_row, column=3, value=_to_number(transaction.sale_price_usd))
        sheet.cell(row=current_row, column=4, value=_to_number(transaction.gross_proceeds_usd))
        sheet.cell(row=current_row, column=5, value=_to_number(transaction.total_fees_usd))
        if transaction.net_proceeds_usd is not None:
            sheet.cell(row=current_row, column=6, value=_to_number(transaction.net_proceeds_usd))
        else:
            sheet.cell(row=current_row, column=6, value=f"=D{current_row}-E{current_row}")
        sheet.cell(row=current_row, column=7, value=transaction.proceeds_currency)
        sheet.cell(row=current_row, column=8, value=_to_number(transaction.converted_net_proceeds))
        current_row += 1

    sheet.cell(row=current_row, column=5, value=f"=SUM(E2:E{current_row - 1})")


def _to_number(value: Decimal | None) -> float | None:
    if value is None:
        return None
    return float(value)


def _parse_date(value: str) -> datetime:
    match = datetime_pattern.search(value)
    normalized = match.group(1) if match else value
    return datetime.strptime(normalized, "%B %d, %Y")


datetime_pattern = __import__("re").compile(r"([A-Z][a-z]+ \d{1,2}, \d{4})")


def _apply_sheet_widths(sheet, width: float, column_count: int) -> None:
    for index in range(1, column_count + 1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def _sum_formula(column: str, rows: list[int]) -> str:
    return "=" + "+".join(f"{column}{row}" for row in rows)


HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
TITLE_FILL = PatternFill(fill_type="solid", fgColor="B8CCE4")
TOTAL_FILL = PatternFill(fill_type="solid", fgColor="EEF3F8")
THIN_SIDE = Side(style="thin", color="A6A6A6")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)


def _style_sales_breakdown(sheet) -> None:
    sheet.freeze_panes = "A2"
    for row in sheet.iter_rows():
        first = row[0]
        values = [cell.value for cell in row]
        if first.coordinate in sheet.merged_cells and first.value:
            for cell in row[:7]:
                cell.fill = TITLE_FILL
                cell.border = THIN_BORDER
        elif values[:7] == [
            "Acquisition Date",
            "Shares Sold",
            "Cost Basis per Share (USD)",
            "Sale Price per Share (USD)",
            "Total Cost Basis (USD)",
            "Total Sale Proceeds (USD)",
            "Gain/Loss (USD)",
        ]:
            for cell in row[:7]:
                cell.fill = HEADER_FILL
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="center")
        elif values[:4] == [
            "Total Shares Sold",
            "Total Cost Basis (USD)",
            "Total Sale Proceeds (USD)",
            "Total Gain/Loss (USD)",
        ]:
            for cell in row[:4]:
                cell.fill = HEADER_FILL
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="center")
                cell.font = Font(bold=True)
        elif row[0].value and isinstance(row[0].value, str) and row[0].value.startswith("=") and row[1].value:
            for cell in row[:4]:
                cell.fill = TOTAL_FILL
                cell.border = THIN_BORDER
                cell.font = Font(bold=True)
        elif row[4].value and isinstance(row[4].value, str) and row[4].value.startswith("=SUM("):
            for cell in row[:7]:
                cell.fill = TOTAL_FILL
                cell.border = THIN_BORDER
                if cell.column >= 5:
                    cell.font = Font(bold=True)
        elif any(cell.value is not None for cell in row[:7]):
            for cell in row[:7]:
                cell.border = THIN_BORDER


def _style_transaction_breakdown(sheet) -> None:
    sheet.freeze_panes = "A2"
    for row_index, row in enumerate(sheet.iter_rows(), start=1):
        if row_index == 1:
            for cell in row[:8]:
                cell.fill = HEADER_FILL
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="center")
        elif row[4].value and isinstance(row[4].value, str) and row[4].value.startswith("=SUM("):
            for cell in row[:8]:
                cell.fill = TOTAL_FILL
                cell.border = THIN_BORDER
                if cell.column >= 5:
                    cell.font = Font(bold=True)
        elif any(cell.value is not None for cell in row[:8]):
            for cell in row[:8]:
                cell.border = THIN_BORDER
